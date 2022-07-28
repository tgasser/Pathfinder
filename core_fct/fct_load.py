"""
Copyright IIASA (International Institute for Applied Systems Analysis), 2019-2022
Contributor(s): Thomas Gasser (gasser@iiasa.ac.at)

This software is a computer program whose purpose is simulate a large number of global climate (and environmental) change scenarios.

This software is governed by the CeCILL license under French law and abiding by the rules of distribution of free software.  You can use, modify and/ or redistribute the software under the terms of the CeCILL license as circulated by CEA, CNRS and INRIA at the following URL "http://www.cecill.info". 

As a counterpart to the access to the source code and rights to copy, modify and redistribute granted by the license, users are provided only with a limited warranty and the software's author, the holder of the economic rights, and the successive licensors have only limited liability. 

In this respect, the user's attention is drawn to the risks associated with loading, using, modifying and/or developing or reproducing the software by the user in light of its specific status of free software, that may mean that it is complicated to manipulate, and that also therefore means that it is reserved for developers and experienced professionals having in-depth computer knowledge. Users are therefore encouraged to load and test the software's suitability as regards their requirements in conditions enabling the security of their systems and/or data to be ensured and,  more generally, to use and operate it in the same conditions as regards security. 

The fact that you are presently reading this means that you have had knowledge of the CeCILL license and that you accept its terms.
"""


import io
import csv
import xlrd
import zipfile
import numpy as np
import pandas as pd
import xarray as xr
import openpyxl as xl

from scipy.stats import theilslopes


##################################################
##   1. LOAD OBSERVATION AND DRIVING DATA
##################################################

## contemporary and historical CO2 budget from Global Carbon Project 
## (Friedlingstein et al., 2022; doi:10.5194/essd-14-1917-2022; https://www.globalcarbonproject.org/carbonbudget/index.htm)
def load_GCB():

    ## properties
    path = 'input_data/GCB/Global_Carbon_Budget_2021v1.0.xlsx'
    var_dic = {'fossil emissions excluding carbonation':'Eff', 'land-use change emissions':'Eluc', 'atmospheric growth':'dCatm', 'ocean sink':'Focean', 'land sink':'Fland', 'cement carbonation sink':'Fcarb', 'budget imbalance':'dCimbal'}

    ## contemporary
    ## load and read
    with open(path, "rb") as f:
        sheet = xl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)['Global Carbon Budget']
    lgd = [[cell.value for cell in row] for row in sheet['A21:H21']][0]
    TMP = np.array([[cell.value for cell in row] for row in sheet['A22:H83']], dtype=float)

    ## format
    var1 = xr.Dataset()
    for n, var in enumerate(lgd[1:]):
        var1[var_dic[var]] = xr.DataArray(TMP[:,n+1][:,np.newaxis], 
            coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'data': ['GCB']}, 
            dims=['year', 'data'], 
            attrs={'units': 'PgC yr-1'})

    ## historical
    ## load and read
    with open(path, "rb") as f:
        sheet = xl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)['Historical Budget']
    lgd = [[cell.value for cell in row] for row in sheet['A16:H16']][0]
    TMP = np.array([[cell.value for cell in row] for row in sheet['A17:H287']], dtype=float)

    ## format
    var2 = xr.Dataset()
    for n, var in enumerate(lgd[1:]):
        var2[var_dic[var]] = xr.DataArray(TMP[:,n+1][:,np.newaxis], 
            coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'data': ['hist']}, 
            dims=['year', 'data'], 
            attrs={'units': 'PgC yr-1'})

    ## return as merged dataset    
    return xr.merge([var1, var2])


## some data from IPCC WR5 WG1 Annex2
## (Prather et al., 2013; doi:10.1017/CBO9781107415324.030; http://www.ipcc.ch/report/ar5/wg1/)
def load_IPCC_AR5_WG1():

    ## properties
    path = 'input_data/AR5/WG1AR5_AIISM_Datafiles.xlsx'
    var_dic = {'CO2':'CO2', 'GHG OTher*':'GHG_nonCO2', 'O3 (Trop)':'O3_trop', 'O3 (Strat)':'O3_strat', 'Aerosol (Total)': 'AER_tot', 'LUC':'LCC_alb', 'H2O (Strat)':'H2O_strat', 'BC Snow':'BC_snow', 'Contrails':'contrails', 'Solar':'solar', 'Volcano':'volc'}

    ## atmospheric concentrations (Table AII.1.1a)
    ## load and read
    with open(path, "rb") as f:
        sheet = xl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)['AII-1-1a']
    lgd = [[cell.value for cell in row] for row in sheet['A8:D8']][0]
    TMP = np.array([[cell.value[:cell.value.find(u'\xb1')] if type(cell.value)==str else cell.value for cell in row] for row in sheet['A10:D107']], dtype=float)

    ## format
    Var1 = xr.Dataset()
    for n, var in enumerate(lgd[1:]):
        Var1[var[:var.find(' ')]] = xr.DataArray(np.interp(np.arange(TMP[0,0], TMP[-1,0]+1), TMP[:,0], TMP[:,n+1]), 
            coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1)}, 
            dims=['year'], 
            attrs={'units': var[var.find('(')+1:var.find(')')]})

    ## radiative forcings (Table AII.1.2)
    ## load and read
    with open(path, "rb") as f:
        sheet = xl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)['AII-1-2']
    lgd = [[cell.value for cell in row] for row in sheet['A8:L8']][0]
    TMP = np.array([[cell.value for cell in row] for row in sheet['A9:L270']], dtype=float)

    ## format
    Var2 = xr.DataArray(TMP[:,1:], 
        coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'forcing': [var_dic[var] for var in lgd[1:]]}, 
        dims=['year', 'forcing'], 
        attrs={'units': 'W m-2'})
    Var2 = xr.Dataset({'RF': Var2})

    ## return
    return xr.merge([Var1, Var2])


## some data from IPCC AR6 WG1 Annex3
## (Dentener et al., 2021; doi:10.1017/9781009157896.017) (Table AIII.3)
## source: https://github.com/chrisroadmap/ar6/blob/main/data_input/observations/LLGHG_history_AR6_v9_for_archive.xlsx
## source: https://github.com/chrisroadmap/ar6/blob/main/data_output/AR6_ERF_1750-2019.csv
def load_IPCC_AR6_WG1():

    ## properties
    path_C = 'input_data/AR6/LLGHG_history_AR6_v9_for_archive.xlsx'
    path_R = 'input_data/AR6/AR6_ERF_1750-2019.csv'
    var_dic = {'co2':'CO2', 'ch4':'CH4', 'n2o':'N2O', 'o3':'O3', 'h2o_stratospheric':'H2O_stratospheric'}

    ## atmospheric concentrations (Table AIII.1a, extended)
    Var1 = xr.Dataset.from_dataframe(pd.read_excel(path_C, sheet_name='mixing_ratios', header=22, index_col=0)).rename({'YYYY': 'year'})
    Var1 = Var1.drop(var for var in Var1 if var not in ['CO2', 'CH4', 'N2O'])
    Var1 = Var1.interp(year=np.arange(Var1.year.min(), Var1.year.max()+1, 1))
    for var in Var1:
        if var in ['CO2']: Var1.attrs['units'] = 'ppm'
        elif var in ['CH4', 'N2O']: Var1.attrs['units'] = 'ppb'
        else: Var1[var].attrs['units'] = 'ppt'

    ## radiative forcings (Table AIII.3, extended)
    Var2 = xr.Dataset.from_dataframe(pd.read_csv(path_R, index_col=0).stack().to_frame()).rename({'level_1': 'forcing', 0: 'ERF'})
    Var2 = Var2.isel(forcing=slice(0,-6)) # remove aggregates
    Var2['forcing'] = [var_dic[var] if var in var_dic.keys() else var for var in Var2.forcing.values]
    Var2['ERF'].attrs['units'] = 'W m-2'

    ## return
    return xr.merge([Var1, Var2])


## recent atmospheric observations from NOAA-ESRL
## (Dlugokencky & Tans; https://www.esrl.noaa.gov/gmd/ccgg/trends)
def load_NOAA_ESRL():

    ## properties
    path_list = ['input_data/obs_CO2/' + path for path in ['co2_annmean_gl.txt', 'co2_annmean_mlo.txt']]
    site_list = ['global', 'mauna_loa']

    ## global average (https://www.esrl.noaa.gov/gmd/ccgg/trends/gl_data.html)
    ## load and read
    with open(path_list[0], 'r') as f:
        TMP = np.array([line for line in csv.reader(f, delimiter=' ', skipinitialspace=True)][57:], dtype=float)

    ## format
    var1 = xr.DataArray(TMP[:,1][:,np.newaxis], 
        coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'site': [site_list[0]]}, 
        dims=['year', 'site'], 
        attrs={'units': 'ppm'})
    var1 = xr.Dataset({'CO2': var1})

    ## Mauna Loa measurements (https://www.esrl.noaa.gov/gmd/ccgg/trends/data.html)
    ## load and read
    with open(path_list[1], 'r') as f:
        TMP = np.array([line for line in csv.reader(f, delimiter=' ', skipinitialspace=True)][57:], dtype=float)

    ## format
    var2 = xr.DataArray(TMP[:,1][:,np.newaxis], 
        coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'site': [site_list[1]]}, 
        dims=['year', 'site'], 
        attrs={'units': 'ppm'})
    var2 = xr.Dataset({'CO2': var2})

    ## return
    return xr.merge([var1, var2])


## global temperature from various sources
## HadCRUT4: (Morice et al., 2012; doi:10.1029/2011JD017187; https://www.metoffice.gov.uk/hadobs/hadcrut4/data/current/download.html)
## HadCRUT5: (Morice et al., 2019; doi:10.1029/2019JD032361; https://www.metoffice.gov.uk/hadobs/hadcrut5/data/current/download.html)
## Cowtan_and_Way: (Cowtan & Way, 2014; doi:10.1002/qj.2297; http://www-users.york.ac.uk/~kdc3/papers/coverage2013/series.html)
## Berkeley_Earth: (Rhodes et al., 2013a,b; doi:10.4172/gigs.1000101, doi:10.4172/gigs.1000103; http://berkeleyearth.org/data/)
## GISTEMP: (Hansen et al., 2010; doi:10.1029/2010RG000345; https://data.giss.nasa.gov/gistemp/)
## NOAAGlobalTemp: (Huang et al., 2020; doi:10.1175/JCLI-D-19-0395.1; https://www.ncei.noaa.gov/products/land-based-station/noaa-global-temp)
## JMA: (Japan Meteorological Agency, 2019; https://ds.data.jma.go.jp/tcc/tcc/products/gwp/gwp.html)
def load_gmst(ref_period=(1850, 1900), ignore_data=[], get_trend=False, trend_over=21, make_plot=False):

    ## properties
    path_list = ['input_data/obs_T/' + path for path in ['HadCRUT.4.6.0.0.annual_ns_avg.txt', 'had4_krig_annual_v2_0_0.txt', 'Land_and_Ocean_summary.txt', 'GLB.Ts+dSST.csv', 'aravg.ann.land_ocean.90S.90N.v5.0.0.202205.asc', 'year_wld.csv']]
    data_list = ['HadCRUT4', 'Cowtan_and_Way', 'Berkeley_Earth', 'GISTEMP', 'NOAAGlobalTemp', 'JMA']

    ## loop on datasets
    var_out = xr.Dataset()
    for n_data in range(len(data_list)):

        ## load and read
        if data_list[n_data] in ['HadCRUT4', 'NOAAGlobalTemp']:
            with open(path_list[n_data], 'r') as f:
                TMP = np.array([line for line in csv.reader(f, delimiter=' ', skipinitialspace=True)][:-1], dtype=float)
        if data_list[n_data] in ['Cowtan_and_Way']:
            with open(path_list[n_data], 'r') as f:
                TMP = np.array([line for line in csv.reader(f, delimiter=' ', skipinitialspace=True)], dtype=float)
        elif data_list[n_data] in ['Berkeley_Earth']:
            with open(path_list[n_data], 'r') as f:
                TMP = np.array([line for line in csv.reader(f, delimiter=' ', skipinitialspace=True)][58:], dtype=float)
        elif data_list[n_data] in ['GISTEMP']:
            with open(path_list[n_data], 'r') as f:
                TMP = np.array([line[:14] for line in csv.reader(f)][2:-1], dtype=float)[:, np.array([0,-1] + list(range(1, 1+12)))]
        elif data_list[n_data] in ['JMA']:
            with open(path_list[n_data], 'r') as f:
                TMP = np.array([line[:2] for line in csv.reader(f)][1:], dtype=float)

        ## format
        var = xr.DataArray(TMP[:,1][:,np.newaxis], 
            coords={'year': np.arange(int(TMP[0,0]), int(TMP[-1,0])+1), 'data': [data_list[n_data]]}, 
            dims=['year', 'data'])
        var_out = xr.merge([var_out, xr.Dataset({'T': var})])

    ## add HadCRUT5 on top (netcdf file)
    with xr.open_dataset('input_data/obs_T/HadCRUT.5.0.1.0.analysis.summary_series.global.annual.nc') as TMP: TMP = TMP.load()
    TMP = TMP.groupby('time.year').mean('time').tas_mean[:-1]
    TMP = TMP.drop(['realization', 'latitude', 'longitude']).assign_coords(data='HadCRUT5').expand_dims('data', -1)
    var_out = xr.merge([var_out, xr.Dataset({'T': TMP})])

    ## remove ignored dataset (with inadequate coverage)
    var_out = var_out.sel(data=[data for data in var_out.data.values if data not in ignore_data])

    ## apply chosen reference period
    if ref_period is not None:
        var_out['T'] = var_out.T - var_out.T.sel(year=slice(*ref_period)).mean('year')

    ## add units
    var_out.T.attrs['units'] = 'K'

    ## calculate various trends
    if get_trend:

        ## trend and smoothed warming
        dT_trend = np.zeros((len(var_out.year)-trend_over, len(var_out.data))) * np.nan
        T_trend = np.zeros((len(var_out.year)-trend_over, len(var_out.data))) * np.nan
        for t in range(len(var_out.year)-trend_over):
            for n_data in range(len(var_out.data)):
                TMP = var_out.T.sel(data=var_out.data[n_data])
                dT_trend[t, n_data] = theilslopes(TMP.values[t:t+trend_over+1])[0]
                T_trend[t, n_data] = np.mean(TMP.values[t:t+trend_over+1]) + 0.5*trend_over*dT_trend[t, n_data]

        ## trend of trend
        ddT_trend = np.zeros((len(var_out.year)-2*trend_over, len(var_out.data))) * np.nan
        for t in range(len(var_out.year)-2*trend_over)[::-1]:
            for n_data in range(len(var_out.data)):      
                ddT_trend[t, n_data] = theilslopes(dT_trend[t:t+trend_over+1, n_data])[0]

        ## add to dataset
        T_trend = xr.DataArray(T_trend, 
            coords={'year': var_out.year[trend_over:], 'data': var_out.data.values}, 
            dims=['year', 'data'], 
            attrs={'units': 'K'})
        dT_trend = xr.DataArray(dT_trend, 
            coords={'year': var_out.year[trend_over:], 'data': var_out.data.values}, 
            dims=['year', 'data'], 
            attrs={'units':'K yr-1'})
        ddT_trend = xr.DataArray(ddT_trend, 
            coords={'year': var_out.year[2*trend_over:], 'data': var_out.data.values}, 
            dims=['year', 'data'], 
            attrs={'units':'K yr-2'})
        var_out = xr.merge([var_out, 
            xr.Dataset({'T_'+str(trend_over)+'yr': T_trend}), 
            xr.Dataset({'dT_'+str(trend_over)+'yr': dT_trend}), 
            xr.Dataset({'ddT_'+str(trend_over)+'yr': ddT_trend})])

    ## optional nice plot
    if get_trend and make_plot:
        import matplotlib.pyplot as plt

        ## get useful values
        T_last = var_out['T_'+str(trend_over)+'yr'][-1]
        yr_last = var_out.year[-1]
        T_over = var_out.T.isel(year=np.arange(-trend_over, 0))
        yr_over = var_out.year.isel(year=np.arange(-trend_over, 0))

        ## make figure and display
        plt.figure()
        for n, data in enumerate(var_out.data.values):
            plt.plot(var_out.year, var_out.T.sel(data=data), color='C'+str(n), lw=1, label=data)
        for n, data in enumerate(var_out.data.values):
            plt.plot(yr_over, T_last.sel(data=data) + (T_last.sel(data=data)  - T_over.mean('year').sel(data=data)) / (yr_last - yr_over.mean('year')) * (yr_over-yr_last), lw=1.5, color='k', ls='--', alpha=0.9)
            plt.plot(yr_over.mean('year'), T_over.mean('year').sel(data=data), color='C'+str(n), ls='none', marker='o', ms=4, mec='k', alpha=0.8)
            plt.plot(yr_last, T_last.sel(data=data), color='C'+str(n), ls='none', marker='o', ms=4, mec='k', alpha=0.8)
        plt.plot(np.nan, np.nan, lw=1.5, color='k', ls='--', label='latest trend')
        plt.plot(ref_period, [0,0], color='k', lw=2, label='reference period')
        plt.legend(loc=2, ncol=1, frameon=False, labelspacing=0.5, fontsize='small')
        plt.xticks(fontsize='small')
        plt.yticks(fontsize='small')
        plt.title(r'{:+.2f}$\pm${:.2f} over {:d}$-${:d}, {:+.2f}$\pm${:.2f} extrapolated to {:d}, w.r.t. {:d}$-${:d}'.format(
            T_over.mean().values, T_over.mean('year').std().values, yr_last.values-trend_over+1, yr_last.values, T_last.mean().values, T_last.std().values, yr_last.values, *ref_period), fontsize='small')
        plt.suptitle(r'global mean surface temperature change (in $\degree$C)', fontsize='medium')
        plt.subplots_adjust(left=0.05, bottom=0.05, right=0.95, top=0.90)
        plt.show()

    ## return
    return var_out


##################################################
##   2. LOAD SCENARIO DATA
##################################################

## RCPs (as defined for CMIP5 in concentration)
## (Meinshausen et al., 2011; doi:10.1007/s10584-011-0156-z; http://www.pik-potsdam.de/~mmalte/rcps/)
def load_RCPs(AR6_ERF=False):

    ## properties
    path = 'input_data/RCPs/'
    path_ar6 = 'input_data/AR6/'
    scen_dic = {'rcp26': 'RCP3PD', 'rcp45': 'RCP45', 'rcp60': 'RCP6', 'rcp85': 'RCP85'}

    ## create empty dataset
    Var = xr.Dataset()
    Var.coords['year'] = np.arange(2005, 2500+1)
    Var.coords['scen'] = np.sort(list(scen_dic.keys()))
    for var in ['CO2', 'ERFx']:
        Var[var] = np.nan * xr.zeros_like(Var.year, dtype=float) * xr.zeros_like(Var.scen, dtype=float)

    ## get original data
    for scen in Var.scen.values:
        
        ## atmospheric concentration
        with xlrd.open_workbook(path + scen_dic[scen] + '_MIDYEAR_CONCENTRATIONS.xls') as wb:
            sheet = wb.sheet_by_name(scen_dic[scen] + '_MIDYEAR_CONCENTRATIONS')
        Var['CO2'].loc[{'scen': scen}] = np.array([cell.value for cell in sheet.col_slice(3, 40-1+2005-1765)])

        ## radiative forcing
        with xlrd.open_workbook(path + scen_dic[scen] + '_MIDYEAR_RADFORCING.xls') as wb:
            sheet = wb.sheet_by_name(scen_dic[scen] + '_MIDYEAR_RADFORCING')
        Sol = np.array([cell.value for cell in sheet.col_slice(3, 61-1+2005-1765)])
        Ant = np.array([cell.value for cell in sheet.col_slice(4, 61-1+2005-1765)])
        CO2 = np.array([cell.value for cell in sheet.col_slice(8, 61-1+2005-1765)])
        Var['ERFx'].loc[{'scen': scen}] = Sol + Ant - CO2

        ## get AR6 estimates of ERFx if requested
        if AR6_ERF:
            TMP = xr.Dataset.from_dataframe(pd.read_csv(path_ar6 + 'ERF_' + scen + '_1750-2500.csv', index_col=0).stack().to_frame()).rename({'level_1': 'forcing', 0: 'ERF'})
            Var['ERFx'].loc[{'scen': scen}] = (TMP.ERF.sel(forcing='solar') + TMP.ERF.sel(forcing='total_anthropogenic') - TMP.ERF.sel(forcing='co2')).sel(year=Var.year)

    ## return
    return Var


## SSPs (as defined for CMIP6 in concentration)
## (Meinshausen et al., 2020; doi:10.5194/gmd-13-3571-2020; https://greenhousegases.science.unimelb.edu.au/)
## (Dentener et al., 2021; doi:10.1017/9781009157896.017) (Table AIII.4 & GitHub)
def load_SSPs():

    ## properties
    path = 'input_data/SSPs/'
    path_ar6 = 'input_data/AR6/'
    scen_dic = {'ssp119': 'T3 - SSP1-1.9 ', 'ssp126': 'T4 -  SSP1-2.6 ', 'ssp245': 'T5 - SSP2-4.5 ', 'ssp370': 'T6 - SSP3-7.0 ', 'ssp370-lowNTCF': 'T7 - SSP3-7.0-lowNTCF ', 
        'ssp434': 'T8 - SSP4-3.4 ', 'ssp460': 'T9 - SSP4-6.0 ', 'ssp534-over': 'T10 - SSP5-3.4-over ', 'ssp585': 'T11 - SSP5-8.5 '}

    ## create empty dataset
    Var = xr.Dataset()
    Var.coords['year'] = np.arange(2014, 2500+1)
    Var.coords['scen'] = np.sort(list(scen_dic.keys()))
    for var in ['CO2', 'ERFx']:
        Var[var] = np.nan * xr.zeros_like(Var.year, dtype=float) * xr.zeros_like(Var.scen, dtype=float)

    ## read zipped data file
    with zipfile.ZipFile(path + 'SUPPLEMENT_DataTables_Meinshausen_6May2020.zip') as f:
        main_file = f.read('SUPPLEMENT_DataTables_Meinshausen_6May2020.xlsx')
    sheet = xl.load_workbook(io.BytesIO(main_file), read_only=True, data_only=True)['T2 - History Year 1750 to 2014']
    CO2_2014 = sheet['B277'].value

    ## get data
    for scen in Var.scen.values:
        
        ## atmospheric concentration
        sheet = xl.load_workbook(io.BytesIO(main_file), read_only=True, data_only=True)[scen_dic[scen]]
        Var['CO2'].loc[{'scen': scen}] = np.append([CO2_2014], np.array([row[0].value for row in sheet['B13:B498']]))

        ## radiative forcing
        TMP = xr.Dataset.from_dataframe(pd.read_csv(path_ar6 + 'ERF_' + scen + '_1750-2500.csv', index_col=0).stack().to_frame()).rename({'level_1': 'forcing', 0: 'ERF'})
        Var['ERFx'].loc[{'scen': scen}] = (TMP.ERF.sel(forcing='solar') + TMP.ERF.sel(forcing='total_anthropogenic') - TMP.ERF.sel(forcing='co2')).sel(year=Var.year)

    ## return
    return Var


##################################################
##   3. LOAD CALIBRATION DATA
##################################################

##=============
## 3.1. CLIMATE
##=============

## load CMIP6 data
def load_CMIP6_climate(get_all_simu=False):

    ## load preprocessed data
    with xr.open_dataset('input_data/CMIP6/climate_CMIP6_for_PathFinder.nc') as TMP:
        var_out = TMP.load()

    ## set to correct units for calibration
    var_out['pr'] *= 1 * 3600 * 24 * 365.25
    var_out['pr'].attrs['units'] =  'mm yr-1'

    ## format output dataset
    if not get_all_simu:
        var_out = var_out.sel(exp=[exp for exp in var_out.exp.values if exp not in ['1pctCO2-cdr', 'abrupt-2xCO2', 'abrupt-0p5xCO2']])
    var_out = var_out.drop(['areacella'])
    var_out = var_out.rename({'exp':'simu', 'mod':'model_clim'})
    var_out = var_out.sel(model_clim=np.sort(var_out.model_clim))
    var_out = var_out.transpose('year', 'simu', 'model_clim')

    ## return
    return var_out


##==================
## 3.2. OCEAN CARBON
##==================

## load CMIP5 data as for OSCAR v2.2
## (Gasser et al., 2017; doi:10.5194/gmd-10-271-2017)
def load_CMIP5_ocean():

    ## CMIP5 metadata
    model_cmip5 = ['bcc-csm1-1', 'CanESM2', 'CESM1-BGC', 'HadGEM2-ES', 'IPSL-CM5A-LR', 'MPI-ESM-LR', 'NorESM1-ME']
    simu_cmip5 = ['piControl', '1pctCO2', 'esmFixClim1', 'esmFdbk1']
    len_cmip5 = 140

    ## initialize
    var_cmip5 = {}
    for var in ['tas', 'tos', 'fgco2', 'dpco2', 'spco2']:
        var_cmip5[var] = np.zeros((len_cmip5, len(simu_cmip5), len(model_cmip5)))

    ## load primary data
    for mod in model_cmip5:
        for sim in simu_cmip5:
            with open('input_data/CMIP5/CMIP5_'+mod+'_'+sim+'.csv', 'r') as f:
                TMP = [line for line in csv.reader(f)]
            var_list = TMP[0]; TMP = np.array(TMP[1:], dtype=float)
            for var in var_list:
                if var in var_cmip5.keys():
                    var_cmip5[var][:, simu_cmip5.index(sim), model_cmip5.index(mod)] = TMP[:len_cmip5, var_list.index(var)]

    ## secondary data for calibration
    ## atmospheric CO2
    var_cmip5['co2'] = np.zeros([len_cmip5, len(simu_cmip5), len(model_cmip5)])
    var_cmip5['co2'][:,1:2+1, :] = 284.32 * 1.01**np.arange(1, len_cmip5+1)[:,np.newaxis, np.newaxis]
    var_cmip5['co2'][:,0::3, :] = 284.32

    ## create output
    ## variables' units
    units = {'tas':'K', 'tos':'K', 'fgco2':'PgC yr-1', 'dpco2':'ppm', 'spco2':'ppm', 'co2':'ppm'}

    ## turned into xarray with coords and attrs
    var_out = {}
    for var in list(var_cmip5.keys()) + ['co2']:
        var_out[var] = xr.DataArray(var_cmip5[var], 
            coords={'year': 1851 + np.arange(len_cmip5), 'simu': simu_cmip5, 'model_ocean': model_cmip5}, 
            dims=['year', 'simu', 'model_ocean'], 
            attrs={'units': units[var]})

    ## return as dataset
    return xr.Dataset(var_out)


## load CMIP6 data
def load_CMIP6_ocean(get_all_simu=False):

    ## load preprocessed data
    with xr.open_dataset('input_data/CMIP6/ocean_CMIP6_for_PathFinder.nc') as TMP:
        var_out = TMP.load()

    ## set to correct units for calibration
    for var in ['spco2', 'dpco2']:
        var_out[var] *= 1E6 / 1013E2
        var_out[var].attrs['units'] =  'ppm'
    var_out['fgco2'] *= var_out.areacello * 3600 * 24 * 365.25 / 1E12
    var_out['fgco2'].attrs['units'] =  'PgC yr-1'
    var_out['tos'] += 273.15
    var_out['tos'].attrs['units'] =  'K'

    ## secondary data for calibration
    ## atmospheric CO2
    var_out['co2'] = xr.zeros_like(var_out.tas) * np.nan
    var_out['co2'].loc[{'exp':'piControl'}] = 284.32
    var_out['co2'].loc[{'exp':'1pctCO2'}] = 284.32 * 1.01**(var_out.year - var_out.year[0] + 1).values[:,np.newaxis]
    var_out['co2'].loc[{'exp':'1pctCO2-bgc'}] = 284.32 * 1.01**(var_out.year - var_out.year[0] + 1).values[:,np.newaxis]
    var_out['co2'].loc[{'exp':'1pctCO2-rad'}] = 284.32
    var_out['co2'].loc[{'exp':'1pctCO2-cdr'}][140-1:280-1] = var_out['co2'].loc[{'exp':'1pctCO2'}][:140][::-1].values
    var_out['co2'].loc[{'exp':'1pctCO2-cdr'}][280-1:] = 284.32
    var_out['co2'].loc[{'exp':'abrupt-4xCO2'}] = 284.32 * 4
    var_out['co2'].attrs['units'] = 'ppm'

    ## format output dataset
    if not get_all_simu:
        var_out = var_out.sel(exp=[exp for exp in var_out.exp.values if exp not in ['1pctCO2-cdr']])
    var_out = var_out.drop(['areacella', 'areacello'])
    var_out = var_out.rename({'exp':'simu', 'mod':'model_ocean'})
    var_out = var_out.sel(model_ocean=np.sort(var_out.model_ocean))
    var_out = var_out.transpose('year', 'simu', 'model_ocean')

    ## return
    return var_out


##====================
## 3.3. LAND CARBON PI
##====================

## load TRENDYv7 data as for OSCAR v3.1
## (Gasser et al., 2020; doi:10.5194/bg-2020-33)
def load_TRENDYv7_land():

    ## TRENDYv7 metadata
    model_trendy = ['CABLE-POP', 'CLASS-CTEM', 'DLEM', 'ISAM', 'JSBACH', 'JULES', 'LPJ', 'LPJ-GUESS', 'LPX-Bern', 'OCN', 'ORCHIDEE', 'ORCHIDEE-CNP', 'SURFEX', 'VISIT']
    len_trendy = 400

    ## initialize
    var_trendy = {}
    for var in ['rh', 'fLitterSoil', 'fFire', 'npp', 'cLitter', 'cVeg', 'fVegLitter', 'fVegSoil', 'cSoil', 'fLateral']:
        var_trendy[var] = np.zeros((len_trendy, len(model_trendy)))

    ## load primary data
    for mod in model_trendy:
        with open('input_data/TRENDYv7/TRENDYv7_'+mod+'_S0.csv', 'r') as f:
            TMP = [line for line in csv.reader(f)]
        var_list = TMP[0]; TMP = np.array(TMP[1:], dtype=float)
        for var in var_list:
            if var in var_trendy.keys():
                var_trendy[var][:, model_trendy.index(mod)] = TMP[:len_trendy, var_list.index(var)]

    ## create output
    ## variables' units
    units = {}
    for var in var_trendy.keys():
        if var in ['cVeg', 'cLitter', 'cSoil']: units[var] = 'PgC'
        else: units[var] = 'PgC yr-1'
    
    ## turned into xarray with coords and attrs
    var_out = {}
    for var in var_trendy.keys():
        var_out[var] = xr.DataArray(var_trendy[var], 
            coords={'year': 1701 + np.arange(len_trendy), 'model_landPI': model_trendy}, 
            dims=['year', 'model_landPI'], 
            attrs={'units': units[var]})

    ## return as dataset
    return xr.Dataset(var_out)


##=================
## 3.4. LAND CARBON
##=================

## load CMIP5 data as for OSCAR v2.2
## (Gasser et al., 2017; doi:10.5194/gmd-10-271-2017)
def load_CMIP5_land():

    ## CMIP5 metadata
    model_cmip5 = ['bcc-csm1-1', 'CanESM2', 'CESM1-BGC', 'HadGEM2-ES', 'IPSL-CM5A-LR', 'MPI-ESM-LR', 'NorESM1-ME']
    simu_cmip5 = ['piControl', '1pctCO2', 'esmFixClim1', 'esmFdbk1']
    len_cmip5 = 140

    ## initialize
    var_cmip5 = {}
    for var in ['npp', 'rh', 'fFire', 'cVeg', 'cLitter', 'cSoil', 'tas']:
        var_cmip5[var] = np.zeros((len_cmip5, len(simu_cmip5), len(model_cmip5)))

    ## load primary data
    for mod in model_cmip5:
        for sim in simu_cmip5:
            with open('input_data/CMIP5/CMIP5_'+mod+'_'+sim+'.csv', 'r') as f:
                TMP = [line for line in csv.reader(f)]
            var_list = TMP[0]; TMP = np.array(TMP[1:], dtype=float)
            for var in var_list:
                if var in var_cmip5.keys():
                    var_cmip5[var][:,simu_cmip5.index(sim), model_cmip5.index(mod)] = TMP[:len_cmip5, var_list.index(var)]

    ## secondary data for calibration
    ## atmospheric CO2
    var_cmip5['co2'] = np.zeros([len_cmip5, len(simu_cmip5), len(model_cmip5)])
    var_cmip5['co2'][:,1:2+1, :] = 284.725 * 1.01**np.arange(1, len_cmip5+1)[:,np.newaxis, np.newaxis]
    var_cmip5['co2'][:,0::3, :] = 284.725

    ## create output
    ## parameters' units
    units = {'npp':'PgC yr-1', 'rh':'PgC yr-1', 'fFire':'PgC yr-1', 'cVeg':'PgC', 'cLitter':'PgC', 'cSoil':'PgC', 'tas':'K', 'co2':'ppm'}

    ## turned into xarray with coords and attrs
    var_out = {}
    for var in list(var_cmip5.keys()) + ['co2']:
        var_out[var] = xr.DataArray(var_cmip5[var], 
            coords={'year': 1851 + np.arange(len_cmip5), 
            'simu': simu_cmip5 , 'model_land': model_cmip5}, 
            dims=['year', 'simu', 'model_land'], 
            attrs={'units': units[var]})

    ## return as dataset
    return xr.Dataset(var_out)


## load CMIP6 data
def load_CMIP6_land(get_all_simu=False):

    ## load preprocessed data
    with xr.open_dataset('input_data/CMIP6/land_CMIP6_for_PathFinder.nc') as TMP:
        var_out = TMP.load()

    ## set to correct units for calibration
    for var in ['cVeg', 'cLitter', 'cSoil', 'cProduct']:
        var_out[var] *= var_out.areacella / 1E12
        var_out[var].attrs['units'] =  'PgC'
    for var in ['npp', 'rh', 'fFire', 'fGrazing', 'fHarvest', 'fLitterSoil', 'fVegLitter', 'fVegSoil']:
        var_out[var] *= var_out.areacella * 3600 * 24 * 365.25 / 1E12
        var_out[var].attrs['units'] =  'PgC yr-1'
    var_out['pr'] *= 1 * 3600 * 24 * 365.25
    var_out['pr'].attrs['units'] =  'mm yr-1'

    ## secondary data for calibration
    ## atmospheric CO2
    var_out['co2'] = xr.zeros_like(var_out.tas) * np.nan
    var_out['co2'].loc[{'exp':'piControl'}] = 284.32
    var_out['co2'].loc[{'exp':'1pctCO2'}] = 284.32 * 1.01**(var_out.year - var_out.year[0] + 1).values[:,np.newaxis]
    var_out['co2'].loc[{'exp':'1pctCO2-bgc'}] = 284.32 * 1.01**(var_out.year - var_out.year[0] + 1).values[:,np.newaxis]
    var_out['co2'].loc[{'exp':'1pctCO2-rad'}] = 284.32
    var_out['co2'].loc[{'exp':'1pctCO2-cdr'}][140-1:280-1] = var_out['co2'].loc[{'exp':'1pctCO2'}][:140][::-1].values
    var_out['co2'].loc[{'exp':'1pctCO2-cdr'}][280-1:] = 284.32
    var_out['co2'].loc[{'exp':'abrupt-4xCO2'}] = 284.32 * 4
    var_out['co2'].attrs['units'] = 'ppm'

    ## format output dataset
    if not get_all_simu:
        var_out = var_out.sel(exp=[exp for exp in var_out.exp.values if exp not in ['1pctCO2-cdr']])
    var_out = var_out.drop(['areacella'])
    var_out = var_out.rename({'exp':'simu', 'mod':'model_land'})
    var_out = var_out.sel(model_land=np.sort(var_out.model_land))
    var_out = var_out.transpose('year', 'simu', 'model_land')

    ## return
    return var_out


##====================
## 3.3. SEA LEVEL RISE
##====================

## load Edwards et al. (2021) data
def load_Edwards_2021_slr(ice):

    ## load preprocessed data
    with xr.open_dataset('input_data/Edwards_2021/slr-'+ice+'_Edwards_2021_for_PathFinder.nc') as TMP:
        var_out = TMP.load()

    ## set to correct units for calibration
    var_out['slr'] *= 10
    var_out['slr'].attrs['units'] =  'mm'

    ## return
    return var_out
    
